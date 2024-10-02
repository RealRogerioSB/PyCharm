from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

# grab the active worksheet
ws = wb.active

# data can be assigned directly to cells
ws["A1"] = 42

# rows can also be appended
ws.append([1, 2, 3])

# Python types automatically be converted
ws["A2"] = datetime.now()

# save the file
wb.save("./src/sample.xlsx")
